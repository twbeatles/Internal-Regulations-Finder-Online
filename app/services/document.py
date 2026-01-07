# -*- coding: utf-8 -*-
import os
import re
import logging
from typing import List, Dict, Tuple, Optional
from collections import Counter
import threading

# 3rd party
# lazy imports in methods to avoid startup errors if missing

from app.utils import logger, FileUtils
from app.config import AppConfig
from app.constants import ErrorMessages, Patterns, Limits

# 커스텀 예외
from app.exceptions import (
    DocumentNotFoundError, DocumentExtractionError, 
    DocumentTypeError, DocumentEmptyError
)

# ============================================================================
# 사전 컴파일된 정규식 패턴 (성능 최적화)
# ============================================================================

# ArticleParser용 패턴
_RE_ARTICLE_SPLIT = re.compile(r'(제\s*\d+\s*조(?:의\s*\d+)?[^\n]*)')
_RE_ARTICLE_MATCH = re.compile(r'제\s*(\d+)\s*조(?:의\s*(\d+))?(.+)?')
_RE_PARAGRAPH_SPLIT = re.compile(r'([①②③④⑤⑥⑦⑧⑨⑩])')
_RE_NUMBER_EXTRACT = re.compile(r'\d+')

# DocumentSplitter용 패턴
_RE_CHAPTER_SPLIT = re.compile(r'(제\s*\d+\s*장[^\n]*)')
_RE_ARTICLE_SPLIT_FULL = re.compile(r'(제\s*\d+\s*조(?:의\s*\d+)?[^\n]*)')

# TextHighlighter용 패턴
_RE_KEYWORD_EXTRACT = re.compile(r'[가-힣]{2,}|[a-zA-Z]{3,}')

# ============================================================================
# 문서 추출기 (v2.0 확장 - Excel, HWP, OCR 지원)
# ============================================================================
class DocumentExtractor:
    def __init__(self):
        self._docx_module = None
        self._pdf_module = None
        self._xlsx_module = None
        self._hwp_module = None
        self._ocr_available = None
    
    @property
    def docx(self):
        if self._docx_module is None:
            try:
                from docx import Document
                self._docx_module = Document
            except ImportError:
                self._docx_module = False
        return self._docx_module
    
    @property
    def pdf(self):
        if self._pdf_module is None:
            try:
                from pypdf import PdfReader
                self._pdf_module = PdfReader
            except ImportError:
                self._pdf_module = False
        return self._pdf_module
    
    @property
    def xlsx(self):
        """Excel 모듈 로드 (v2.0)"""
        if self._xlsx_module is None:
            try:
                import openpyxl
                self._xlsx_module = openpyxl
            except ImportError:
                self._xlsx_module = False
        return self._xlsx_module
    
    @property
    def hwp(self):
        """HWP 모듈 로드 (v2.0) - olefile 기반 기본 추출"""
        if self._hwp_module is None:
            try:
                import olefile
                self._hwp_module = olefile
            except ImportError:
                self._hwp_module = False
        return self._hwp_module
    
    @property
    def ocr_available(self):
        """OCR 가용성 확인 (v2.0)"""
        if self._ocr_available is None:
            try:
                import pytesseract
                from PIL import Image
                # Tesseract 설치 확인
                pytesseract.get_tesseract_version()
                self._ocr_available = True
            except Exception:
                self._ocr_available = False
        return self._ocr_available
    
    def extract(self, path: str) -> Tuple[str, Optional[str]]:
        """문서에서 텍스트 추출
        
        Args:
            path: 추출할 파일 경로
            
        Returns:
            Tuple[str, Optional[str]]: (추출된 텍스트, 에러 메시지 또는 None)
            
        Note:
            에러 발생 시에도 예외를 발생시키지 않고 (빈 문자열, 에러 메시지) 반환.
            이는 일괄 처리 시 하나의 실패가 전체를 중단시키지 않도록 하기 위함.
        """
        if not path:
            return "", ErrorMessages.FILE_NOT_FOUND
        
        if not os.path.exists(path):
            logger.debug(f"파일 없음: {path}")
            return "", f"{ErrorMessages.FILE_NOT_FOUND}: {path}"
        
        if not os.path.isfile(path):
            return "", f"파일이 아님: {path}"
        
        ext = os.path.splitext(path)[1].lower()
        
        # 지원 형식 확인
        if ext not in AppConfig.SUPPORTED_EXTENSIONS:
            return "", f"{ErrorMessages.FILE_TYPE_NOT_SUPPORTED}: {ext}"
        
        # 파일 형식별 추출
        try:
            if ext == '.txt':
                return self._extract_txt(path)
            elif ext == '.docx':
                return self._extract_docx(path)
            elif ext == '.pdf':
                return self._extract_pdf(path)
            elif ext in ['.xlsx', '.xls']:
                return self._extract_xlsx(path)
            elif ext == '.hwp':
                return self._extract_hwp(path)
            else:
                return "", f"{ErrorMessages.FILE_TYPE_NOT_SUPPORTED}: {ext}"
        except Exception as e:
            logger.error(f"문서 추출 중 예상치 못한 오류: {path} - {e}")
            return "", f"추출 오류: {str(e)}"
    
    def _extract_txt(self, path: str) -> Tuple[str, Optional[str]]:
        return FileUtils.safe_read(path)
    
    def _extract_docx(self, path: str) -> Tuple[str, Optional[str]]:
        if not self.docx:
            return "", "DOCX 라이브러리 없음 (pip install python-docx)"
        try:
            doc = self.docx(path)
            parts = []
            for para in doc.paragraphs:
                if para.text.strip():
                    parts.append(para.text.strip())
            for table in doc.tables:
                for row in table.rows:
                    cells = [c.text.strip() for c in row.cells if c.text.strip()]
                    if cells:
                        parts.append(' | '.join(cells))
            return '\n\n'.join(parts), None
        except Exception as e:
            return "", f"DOCX 오류: {e}"
    
    def _extract_pdf(self, path: str) -> Tuple[str, Optional[str]]:
        if not self.pdf:
            return "", "PDF 라이브러리 없음 (pip install pypdf)"
        try:
            reader = self.pdf(path)
            if reader.is_encrypted:
                try:
                    reader.decrypt('')
                except Exception:
                    return "", "암호화된 PDF"
            texts = []
            for page in reader.pages:
                try:
                    text = page.extract_text()
                    if text and text.strip():
                        texts.append(text.strip())
                except Exception:
                    continue
            
            # 텍스트가 없으면 OCR 시도 (v2.0)
            if not texts:
                if self.ocr_available:
                    return self._extract_pdf_ocr(path)
                return "", "텍스트 없음 (이미지 PDF - OCR 미설치)"
            return '\n\n'.join(texts), None
        except Exception as e:
            return "", f"PDF 오류: {e}"
    
    def _extract_pdf_ocr(self, path: str) -> Tuple[str, Optional[str]]:
        """OCR을 사용한 이미지 PDF 텍스트 추출 (v2.0)"""
        try:
            import pytesseract
            from pdf2image import convert_from_path
            
            # PDF를 이미지로 변환
            images = convert_from_path(path, dpi=150)
            texts = []
            
            for i, image in enumerate(images):
                try:
                    # 한국어 + 영어 OCR
                    text = pytesseract.image_to_string(image, lang='kor+eng')
                    if text.strip():
                        texts.append(f"[페이지 {i+1}]\n{text.strip()}")
                except Exception as e:
                    logger.warning(f"OCR 페이지 {i+1} 실패: {e}")
                    continue
            
            if texts:
                return '\n\n'.join(texts), None
            return "", "OCR 텍스트 추출 실패"
        except Exception as e:
            return "", f"OCR 오류: {e}"
    
    def _extract_xlsx(self, path: str) -> Tuple[str, Optional[str]]:
        """Excel 파일 텍스트 추출 (v2.0)"""
        if not self.xlsx:
            return "", "Excel 라이브러리 없음 (pip install openpyxl)"
        try:
            wb = self.xlsx.load_workbook(path, read_only=True, data_only=True)
            texts = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_texts = [f"[시트: {sheet_name}]"]
                
                for row in sheet.iter_rows():
                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            row_values.append(str(cell.value).strip())
                    if any(row_values):
                        sheet_texts.append(' | '.join(filter(None, row_values)))
                
                if len(sheet_texts) > 1:  # 시트 제목만 있으면 스킵
                    texts.append('\n'.join(sheet_texts))
            
            wb.close()
            return '\n\n'.join(texts), None
        except Exception as e:
            return "", f"Excel 오류: {e}"
    
    def _extract_hwp(self, path: str) -> Tuple[str, Optional[str]]:
        """HWP 파일 텍스트 추출 (v2.0) - olefile 기반 기본 추출
        
        리소스 관리: try-finally로 OLE 파일 핸들 확실히 닫기
        """
        if not self.hwp:
            return "", "HWP 라이브러리 없음 (pip install olefile)"
        
        ole = None
        try:
            ole = self.hwp.OleFileIO(path)
            texts = []
            
            # PrvText 스트림에서 텍스트 추출 시도 (미리보기 텍스트)
            if ole.exists('PrvText'):
                try:
                    prv_text = ole.openstream('PrvText').read()
                    decoded = prv_text.decode('utf-16le', errors='ignore')
                    decoded = decoded.replace('\x00', '')
                    if decoded.strip():
                        texts.append(decoded.strip())
                except Exception as e:
                    logger.debug(f"HWP PrvText 추출 실패: {e}")
            
            # BodyText 섹션에서 텍스트 추출 시도
            for entry in ole.listdir():
                entry_path = '/'.join(entry)
                if 'BodyText' in entry_path or 'Section' in entry_path:
                    try:
                        stream = ole.openstream(entry)
                        data = stream.read()
                        for encoding in ['utf-16le', 'cp949', 'utf-8']:
                            try:
                                decoded = data.decode(encoding, errors='ignore')
                                decoded = decoded.replace('\x00', '')
                                if decoded.strip() and len(decoded.strip()) > 10:
                                    texts.append(decoded.strip())
                                break
                            except Exception:
                                continue
                    except Exception as e:
                        logger.debug(f"HWP BodyText 섹션 읽기 실패: {entry_path} - {e}")
                        continue
            
            if texts:
                unique_texts = list(dict.fromkeys(texts))
                return '\n\n'.join(unique_texts), None
            return "", "HWP 텍스트 추출 실패 (빈 파일 또는 지원되지 않는 형식)"
        except Exception as e:
            logger.warning(f"HWP 파일 처리 오류: {path} - {e}")
            return "", f"HWP 오류: {e}"
        finally:
            # 리소스 정리: 항상 OLE 파일 닫기
            if ole is not None:
                try:
                    ole.close()
                except Exception as e:
                    logger.debug(f"HWP OLE 파일 닫기 실패: {e}")

# ============================================================================
# 조문/조항 파서 (v2.0)
# ============================================================================
class ArticleParser:
    """규정 문서의 조문 구조 파싱
    
    성능 최적화:
    - 사전 컴파일된 정규식 사용 (모듈 레벨에 정의)
    - 반복 패턴 매칭 최소화
    """
    
    # 클래스 레벨 패턴 (사전 컴파일된 모듈 레벨 패턴 참조)
    ARTICLE_PATTERNS = [
        (re.compile(r'제\s*(\d+)\s*장[^\n]*'), 'chapter'),
        (re.compile(r'제\s*(\d+)\s*절[^\n]*'), 'section'),
        (re.compile(r'제\s*(\d+)\s*조[^\n]*'), 'article'),
        (re.compile(r'제\s*(\d+)\s*조의\s*(\d+)[^\n]*'), 'article_sub'),
    ]
    
    ITEM_PATTERNS = [
        (re.compile(r'①|②|③|④|⑤|⑥|⑦|⑧|⑨|⑩'), 'paragraph'),
        (re.compile(r'^\s*(\d+)\.\s*'), 'numbered'),
        (re.compile(r'^\s*[가-하]\.\s*'), 'korean'),
    ]
    
    def parse_articles(self, content: str) -> List[Dict]:
        """조문별로 분리된 구조 반환 (사전 컴파일된 패턴 사용)"""
        articles = []
        # 사전 컴파일된 패턴 사용
        parts = _RE_ARTICLE_SPLIT.split(content)
        
        current_article = None
        for part in parts:
            match = _RE_ARTICLE_MATCH.match(part.strip())
            if match:
                if current_article:
                    articles.append(current_article)
                
                article_num = match.group(1)
                sub_num = match.group(2) or ""
                title = match.group(3) or ""
                
                current_article = {
                    'number': f"제{article_num}조" + (f"의{sub_num}" if sub_num else ""),
                    'title': title.strip().strip('()（）[]'),
                    'content': "",
                    'paragraphs': []
                }
            elif current_article and part.strip():
                current_article['content'] += part
                
                # 사전 컴파일된 패턴으로 항 분리
                para_split = _RE_PARAGRAPH_SPLIT.split(part)
                for i in range(1, len(para_split), 2):
                    if i+1 < len(para_split):
                        current_article['paragraphs'].append({
                            'marker': para_split[i],
                            'content': para_split[i+1].strip()
                        })
        
        if current_article:
            articles.append(current_article)
        
        return articles
    
    def search_article(self, articles: List[Dict], query: str) -> List[Dict]:
        """조문에서 검색"""
        results = []
        query_lower = query.lower()
        
        for article in articles:
            score = 0
            if query_lower in article.get('title', '').lower():
                score += 3
            if query_lower in article.get('content', '').lower():
                score += 1
            if query_lower in article.get('number', '').lower():
                score += 5
            
            if score > 0:
                results.append({**article, 'score': score})
        
        return sorted(results, key=lambda x: x['score'], reverse=True)
    
    def get_article_by_number(self, articles: List[Dict], number: str) -> Optional[Dict]:
        """조문 번호로 조문 찾기 (사전 컴파일된 패턴 사용)"""
        num_match = _RE_NUMBER_EXTRACT.search(number)
        if not num_match:
            return None
        
        target_num = num_match.group()
        for article in articles:
            if target_num in article.get('number', ''):
                return article
        return None

# ============================================================================
# 문서 분할기 (v2.0)
# ============================================================================
class DocumentSplitter:
    """대용량 규정집을 개별 규정 파일로 분할
    
    Usage:
        splitter = DocumentSplitter(chunk_size=500, chunk_overlap=50)
        chunks = splitter.split(text)
    """
    
    def __init__(self, chunk_size: int = 500, chunk_overlap: int = 50):
        """
        Args:
            chunk_size: 각 청크의 최대 문자 수
            chunk_overlap: 청크 간 겹치는 문자 수
        """
        self.chunk_size = max(chunk_size, Limits.MIN_CHUNK_SIZE)
        self.chunk_overlap = min(chunk_overlap, chunk_size // 2)
    
    def split(self, text: str) -> List[str]:
        """텍스트를 청크로 분할
        
        Args:
            text: 분할할 전체 텍스트
            
        Returns:
            분할된 청크 리스트
        """
        if not text or not text.strip():
            return []
        
        # 문단 단위로 먼저 분리
        paragraphs = text.split('\n\n')
        
        chunks = []
        current_chunk = ""
        
        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
            
            # 현재 청크 + 새 문단이 chunk_size 이하면 추가
            if len(current_chunk) + len(para) + 2 <= self.chunk_size:
                current_chunk += ('\n\n' + para) if current_chunk else para
            else:
                # 현재 청크 저장
                if current_chunk:
                    chunks.append(current_chunk)
                
                # 새 문단이 chunk_size보다 크면 강제 분할
                if len(para) > self.chunk_size:
                    for i in range(0, len(para), self.chunk_size - self.chunk_overlap):
                        chunk = para[i:i + self.chunk_size]
                        if chunk.strip():
                            chunks.append(chunk)
                    current_chunk = ""
                else:
                    # overlap 적용
                    if chunks and self.chunk_overlap > 0:
                        overlap_text = chunks[-1][-self.chunk_overlap:]
                        current_chunk = overlap_text + '\n\n' + para
                    else:
                        current_chunk = para
        
        # 마지막 청크 저장
        if current_chunk.strip():
            chunks.append(current_chunk)
        
        return chunks
    
    def split_by_chapters(self, content: str, filename: str = "") -> List[Dict]:
        pattern = r'(제\s*\d+\s*장[^\n]*)'
        return self._split_by_pattern(content, pattern, 'chapter', filename)
    
    def split_by_articles(self, content: str, filename: str = "") -> List[Dict]:
        pattern = r'(제\s*\d+\s*조(?:의\s*\d+)?[^\n]*)'
        return self._split_by_pattern(content, pattern, 'article', filename)
    
    def _split_by_pattern(self, content: str, pattern: str, split_type: str, filename: str) -> List[Dict]:
        parts = re.split(pattern, content)
        results = []
        
        current_title = None
        current_content = ""
        
        for part in parts:
            if re.match(pattern, part):
                if current_title:
                    results.append({
                        'title': current_title.strip(),
                        'content': current_content.strip(),
                        'type': split_type,
                        'source': filename
                    })
                current_title = part
                current_content = ""
            else:
                current_content += part
        
        if current_title:
            results.append({
                'title': current_title.strip(),
                'content': current_content.strip(),
                'type': split_type,
                'source': filename
            })
        
        return results
    
    def split_by_size(self, content: str, max_size: int = 5000) -> List[Dict]:
        results = []
        paragraphs = content.split('\n\n')
        
        current_chunk = ""
        chunk_num = 1
        
        for para in paragraphs:
            if len(current_chunk) + len(para) > max_size:
                if current_chunk:
                    results.append({
                        'title': f'파트 {chunk_num}',
                        'content': current_chunk.strip(),
                        'type': 'size_split'
                    })
                    chunk_num += 1
                    current_chunk = para
                else:
                    results.append({
                        'title': f'파트 {chunk_num}',
                        'content': para[:max_size],
                        'type': 'size_split'
                    })
                    chunk_num += 1
            else:
                current_chunk += '\n\n' + para if current_chunk else para
        
        if current_chunk:
            results.append({
                'title': f'파트 {chunk_num}',
                'content': current_chunk.strip(),
                'type': 'size_split'
            })
        
        return results

# ============================================================================
# 문서 비교기 (v2.0)
# ============================================================================
class DocumentComparator:
    """두 문서 간 차이점 비교"""
    
    def compare(self, doc1: str, doc2: str) -> Dict:
        import difflib
        
        lines1 = doc1.splitlines(keepends=True)
        lines2 = doc2.splitlines(keepends=True)
        
        differ = difflib.unified_diff(lines1, lines2, lineterm='')
        diff_lines = list(differ)
        
        added = sum(1 for line in diff_lines if line.startswith('+') and not line.startswith('+++'))
        removed = sum(1 for line in diff_lines if line.startswith('-') and not line.startswith('---'))
        
        html_diff = difflib.HtmlDiff()
        html_table = html_diff.make_table(lines1, lines2, context=True)
        
        return {
            'added_lines': added,
            'removed_lines': removed,
            'total_changes': added + removed,
            'diff_text': ''.join(diff_lines),
            'diff_html': html_table,
            'similarity': difflib.SequenceMatcher(None, doc1, doc2).ratio()
        }
    
    def highlight_changes(self, doc1: str, doc2: str) -> Tuple[str, str]:
        import difflib
        
        matcher = difflib.SequenceMatcher(None, doc1, doc2)
        
        result1 = []
        result2 = []
        
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'equal':
                result1.append(doc1[i1:i2])
                result2.append(doc2[j1:j2])
            elif tag == 'delete':
                result1.append(f'<del>{doc1[i1:i2]}</del>')
            elif tag == 'insert':
                result2.append(f'<ins>{doc2[j1:j2]}</ins>')
            elif tag == 'replace':
                result1.append(f'<del>{doc1[i1:i2]}</del>')
                result2.append(f'<ins>{doc2[j1:j2]}</ins>')
        
        return ''.join(result1), ''.join(result2)

# ============================================================================
# 텍스트 하이라이터
# ============================================================================
class TextHighlighter:
    @staticmethod
    def highlight(text: str, query: str, tag: str = 'mark') -> str:
        if not text or not query:
            return text
        
        keywords = [kw.strip() for kw in query.split() if len(kw.strip()) >= 2]
        if not keywords:
            return text
        
        result = text
        for keyword in keywords:
            pattern = re.compile(re.escape(keyword), re.IGNORECASE)
            result = pattern.sub(f'<{tag}>\\g<0></{tag}>', result)
        
        return result
    
    @staticmethod
    def extract_keywords(documents: List[str], top_k: int = 50) -> List[str]:
        if not documents:
            return []
        
        word_freq = Counter()
        for doc in documents:
            # 사전 컴파일된 패턴 사용 (성능 최적화)
            words = _RE_KEYWORD_EXTRACT.findall(doc)
            word_freq.update(words)
        
        stopwords = {'있는', '하는', '및', '등', '이', '가', '을', '를', '의', '에', '로', '으로'}
        keywords = [w for w, _ in word_freq.most_common(top_k * 2) if w not in stopwords]
        
        return keywords[:top_k]
